from openpyxl import load_workbook
import json
from collections import defaultdict

def excel_to_json_openpyxl(file_path):
    wb = load_workbook(filename=file_path, data_only=True)
    
    result = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        merged_cell_map = {}
        for merged_range in sheet.merged_cells.ranges:
            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            value = top_left_cell.value
            
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cell_map[(row, col)] = value
        
        headers = []
        max_cols = sheet.max_column
        
        main_headers = {}
        sub_headers = {}
        
        current_main_header = None
        for col_num in range(1, max_cols + 1):
            cell_value = sheet.cell(1, col_num).value
            
            if (1, col_num) in merged_cell_map:
                cell_value = merged_cell_map[(1, col_num)]
            
            if cell_value and str(cell_value).strip():
                current_main_header = str(cell_value).strip()
            
            main_headers[col_num] = current_main_header
        
        # Process sub headers (row 2)
        for col_num in range(1, max_cols + 1):
            cell_value = sheet.cell(2, col_num).value
            
            if (2, col_num) in merged_cell_map:
                cell_value = merged_cell_map[(2, col_num)]
            
            if cell_value and str(cell_value).strip():
                sub_headers[col_num] = str(cell_value).strip()
            else:
                sub_headers[col_num] = None
        
        for col_num in range(1, max_cols + 1):
            main_header = main_headers.get(col_num)
            sub_header = sub_headers.get(col_num)
            
            main_header = main_header.strip() if main_header and str(main_header).strip() else None
            sub_header = sub_header.strip() if sub_header and str(sub_header).strip() else None
            
            if main_header and sub_header and main_header != sub_header:
                header_name = f"{main_header} | {sub_header}"
            elif main_header and not sub_header:
                header_name = f"{main_header}@SINGLE"
            elif main_header:
                header_name = f"{main_header}@SINGLE"
            elif sub_header:
                header_name = f"{sub_header}@SINGLE"
            else:
                header_name = f"Column_{col_num}@SINGLE"
            
            headers.append(header_name)
        
        # Process data rows
        data = []
        for row_num in range(3, sheet.max_row + 1):  # Start from row 3 (after headers)
            row_data = {}
            has_data = False
            
            for col_num in range(1, min(len(headers) + 1, max_cols + 1)):
                cell_value = sheet.cell(row_num, col_num).value
                
                if (row_num, col_num) in merged_cell_map:
                    cell_value = merged_cell_map[(row_num, col_num)]
                
                if cell_value is not None:
                    has_data = True
                
                if col_num <= len(headers):
                    row_data[headers[col_num - 1]] = cell_value
            
            if has_data:
                data.append(row_data)
        
        result[sheet_name] = {
            'headers': headers,
            'data': data
        }
    
    return result

def excel_to_json_simple(file_path):
    wb = load_workbook(filename=file_path, data_only=True)
    
    result = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        headers = []
        for col in range(1, sheet.max_column + 1):
            main_header = sheet.cell(1, col).value
            sub_header = sheet.cell(2, col).value
            
            if main_header and sub_header:
                headers.append(f"{main_header} | {sub_header}")
            elif main_header:
                headers.append(main_header)
            else:
                headers.append(sub_header)
        
        for row_idx in range(3, sheet.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                if col_idx <= sheet.max_column:
                    value = sheet.cell(row_idx, col_idx).value
                    row_data[header] = value if value is not None else None
            
            if any(value is not None for value in row_data.values()):
                data.append(row_data)
        
        result[sheet_name] = data
    
    return result

try:
    json_data = excel_to_json_openpyxl('output_summary_latest.xlsx')
except Exception as e:
    print(f"Complex parser failed: {e}, trying simple parser...")
    json_data = excel_to_json_simple('output_summary_latest.xlsx')

with open('output_summary_latest.json', 'w', encoding='utf-8') as f:
    json.dump(json_data, f, indent=4, ensure_ascii=False)

print("Successfully created output_summary_latest.json")